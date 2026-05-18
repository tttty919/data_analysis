"""Report exporters - HTML and Excel format."""
import json
from pathlib import Path
from typing import Dict, Any
from backend.config import TASKS_DIR


def generate_html_report(task_id: str, report: Dict[str, Any],
                         base_dir: Path = None) -> Path:
    task_dir = TASKS_DIR / task_id
    task_dir.mkdir(parents=True, exist_ok=True)

    executive_summary = report.get("executive_summary", "暂无摘要")
    key_insights = report.get("key_insights", [])
    recommendations = report.get("recommendations", [])

    # Generate insight cards HTML
    insight_html = ""
    for ins in key_insights:
        insight_html += f'<div class="insight-item">• {ins}</div>'

    # Generate recommendation cards HTML
    rec_html = ""
    for rec in recommendations:
        rec_html += f"""
        <div class="rec-card">
            <div class="rec-title">{rec.get('title', '')}</div>
            <div class="rec-detail">{rec.get('detail', '')}</div>
            <div class="rec-action">➤ {rec.get('action', '')}</div>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>数据分析报告 - {task_id}</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"></script>
<style>
:root {{ --c-bg:#f3f4f6; --c-surface:#fff; --c-border:#e5e7eb; --c-text:#1f2937; --c-text2:#6b7280; --c-accent:#6366f1; --c-green:#22c55e; --c-red:#ef4444; --radius:10px; }}
* {{ box-sizing:border-box; margin:0; padding:0; }}
body {{ font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif; background:var(--c-bg); color:var(--c-text); padding:24px; }}
.container {{ max-width:1100px; margin:0 auto; }}
h1 {{ font-size:24px; margin-bottom:8px; }}
.meta {{ color:var(--c-text2); font-size:13px; margin-bottom:24px; }}
h2 {{ font-size:18px; margin:24px 0 12px; padding-bottom:8px; border-bottom:2px solid var(--c-accent); }}
.summary {{ background:var(--c-surface); padding:20px; border-radius:var(--radius); line-height:1.8; white-space:pre-wrap; margin-bottom:24px; box-shadow:0 1px 3px rgba(0,0,0,.06); }}
.insight-item {{ background:var(--c-surface); padding:12px 16px; border-radius:8px; margin-bottom:8px; border-left:3px solid var(--c-accent); line-height:1.6; }}
.rec-card {{ background:var(--c-surface); padding:20px; border-radius:var(--radius); margin-bottom:12px; box-shadow:0 1px 3px rgba(0,0,0,.06); }}
.rec-title {{ font-size:16px; font-weight:600; color:var(--c-accent); margin-bottom:8px; }}
.rec-detail {{ line-height:1.7; margin-bottom:8px; }}
.rec-action {{ background:#eef2ff; padding:10px 14px; border-radius:6px; color:#4338ca; font-size:14px; }}
.chart-row {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:16px; }}
.chart-box {{ background:var(--c-surface); padding:16px; border-radius:var(--radius); box-shadow:0 1px 3px rgba(0,0,0,.06); height:400px; }}
@media print {{ body {{ background:#fff; }} }}
</style>
</head>
<body>
<div class="container">
<h1>📊 智能数据分析报告</h1>
<div class="meta">任务ID: {task_id} | 生成时间: {report.get('metadata', {}).get('generated_at', '')}</div>

<h2>📋 执行摘要</h2>
<div class="summary">{executive_summary}</div>

<h2>💡 关键洞察</h2>
{insight_html if insight_html else '<p style="color:var(--c-text2)">暂无洞察数据</p>'}

<h2>🎯 行动建议</h2>
{rec_html if rec_html else '<p style="color:var(--c-text2)">暂无建议</p>'}

<h2>📈 可视化图表</h2>
<div class="chart-row">
    <div class="chart-box" id="chart-dist"></div>
    <div class="chart-box" id="chart-corr"></div>
</div>
</div>
<script>
(function() {{
    var dist = document.getElementById('chart-dist');
    if (dist) {{
        var c = echarts.init(dist);
        c.setOption({{ title:{{text:'数据概览'}}, xAxis:{{type:'category',data:[]}}, yAxis:{{type:'value'}}, series:[{{data:[],type:'bar'}}] }});
    }}
}})();
</script>
</body>
</html>"""

    report_path = task_dir / "report.html"
    report_path.write_text(html, encoding="utf-8")
    return report_path


def generate_excel_report(task_id: str, report: Dict[str, Any],
                          eda_results: Dict[str, Any] = None) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    task_dir = TASKS_DIR / task_id
    task_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # Sheet 1: Overview
    ws1 = wb.active
    ws1.title = "概览"
    ws1.append(["智能数据分析报告"])
    ws1.append(["任务ID", task_id])
    ws1.append(["生成时间", report.get("metadata", {}).get("generated_at", "")])
    ws1.append(["摘要", report.get("executive_summary", "")])

    # Sheet 2: Key Insights
    ws2 = wb.create_sheet("关键洞察")
    ws2.append(["序号", "洞察内容"])
    for i, ins in enumerate(report.get("key_insights", []), 1):
        ws2.append([i, ins])

    # Sheet 3: Statistics (if EDA available)
    if eda_results:
        ws3 = wb.create_sheet("描述性统计")
        stats = eda_results.get("statistical_analysis", {})
        if stats:
            cols = ["mean", "std", "min", "q25", "median", "q75", "max", "skewness"]
            ws3.append(["列名"] + cols)
            for col_name, s in stats.items():
                ws3.append([col_name] + [s.get(c, "") for c in cols])

        # Sheet 4: Outliers
        ws4 = wb.create_sheet("异常值检测")
        outliers = eda_results.get("outlier_analysis", {})
        ws4.append(["列名", "异常值数量", "异常值占比(%)", "下界", "上界"])
        for col_name, o in outliers.items():
            ws4.append([col_name, o.get("outlier_count", ""), o.get("outlier_percentage", ""),
                        o.get("lower_bound", ""), o.get("upper_bound", "")])

    # Sheet 5: Recommendations
    ws5 = wb.create_sheet("行动建议")
    ws5.append(["序号", "标题", "详情", "具体行动"])
    for i, rec in enumerate(report.get("recommendations", []), 1):
        ws5.append([i, rec.get("title", ""), rec.get("detail", ""), rec.get("action", "")])

    # Styling
    for ws in [ws2, ws5]:
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 60

    report_path = task_dir / "report.xlsx"
    wb.save(str(report_path))
    return report_path
